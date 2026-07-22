#!/usr/bin/env python3
import argparse
import csv
import os
import re
import subprocess
import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parents[1]
REQUIRED_FM = {"title","document_id","status","owner_role","reviewers","version","last_reviewed","next_review","programme","sensitivity","content_layer","source_references"}
ALLOWED_STATUS = {"draft","review-required","approved","deprecated","historical-example"}
RAW_SOURCE_PREFIXES = ("Journey - Boys - Paul/","Journey - Paul/","Journey - Richard/",".work/")
RAW_SOURCE_FILES = {"ROP - who does what.xlsx","TJ26-1 ROP (Weekend Program) _ April 2026.xlsx"}
TEXT_SUFFIXES = {".md",".csv",".json",".yml",".yaml",".txt",".py",".html",".css",".js",".toml",".ini",".cfg",".sh",".mk",".Makefile"}

def parse_fm(raw):
    data = {}
    for line in raw.splitlines():
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        value = value.strip()
        if value == "[]":
            parsed = []
        elif value.startswith("[") and value.endswith("]"):
            parsed = [v.strip().strip('\"') for v in value[1:-1].split(",") if v.strip()]
        else:
            parsed = value.strip('\"')
        data[key.strip()] = parsed
    return data

def tracked_files():
    return subprocess.check_output(["git","ls-files"], cwd=ROOT, text=True).splitlines()

def is_text(path):
    return path.suffix in TEXT_SUFFIXES or path.name == "Makefile"

def add(errors, msg):
    errors.append(msg)

def check_tracked(errors, files):
    for p in files:
        if p.startswith(RAW_SOURCE_PREFIXES) or p in RAW_SOURCE_FILES:
            add(errors, f"raw or ignored source tracked: {p}")
        path = ROOT / p
        if path.exists() and path.is_file() and path.stat().st_size > 50 * 1024 * 1024:
            add(errors, f"tracked file larger than 50MB: {p}")

def check_privacy(errors, files):
    email = re.compile(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}")
    secret = re.compile(r"(?i)(api[_-]?key|secret|token|password)\s*[:=]\s*[A-Za-z0-9_\-]{12,}")
    phone = re.compile(r"(?<![A-Fa-f0-9])(?:\+?\d[\d .()/-]{8,}\d)(?![A-Fa-f0-9])")
    for p in files:
        path = ROOT / p
        if not path.exists() or not path.is_file() or not is_text(path):
            continue
        text = path.read_text(encoding="utf-8", errors="ignore")
        scrub = "\n".join(line for line in text.splitlines() if "sha256" not in line.lower() and not re.search(r"[0-9a-f]{32,}", line))
        if email.search(scrub) and "example.invalid" not in scrub:
            add(errors, f"possible email in {p}")
        if secret.search(scrub):
            add(errors, f"possible secret in {p}")
        if p.startswith("archive-register/"):
            continue
        for match in phone.finditer(scrub):
            digits = re.sub(r"\D", "", match.group(0))
            if len(digits) >= 10 and not digits.startswith(("2024","2025","2026")):
                add(errors, f"possible private number in {p}")
                break

def check_front_matter(errors):
    ids = {}
    for base in [ROOT / "docs", ROOT / "templates"]:
        for path in base.rglob("*.md"):
            text = path.read_text(encoding="utf-8")
            rel = path.relative_to(ROOT)
            if not text.startswith("---\n"):
                add(errors, f"{rel} missing front matter")
                continue
            end = text.find("\n---\n", 4)
            if end == -1:
                add(errors, f"{rel} malformed front matter")
                continue
            data = parse_fm(text[4:end])
            missing = REQUIRED_FM - set(data)
            if missing:
                add(errors, f"{rel} missing front matter keys {sorted(missing)}")
            if data.get("status") not in ALLOWED_STATUS:
                add(errors, f"{rel} invalid status {data.get('status')}")
            if data.get("status") == "approved":
                add(errors, f"{rel} is approved without final human approval")
            doc_id = data.get("document_id")
            if doc_id:
                if doc_id in ids:
                    add(errors, f"duplicate document_id {doc_id}: {rel} and {ids[doc_id]}")
                ids[doc_id] = rel
            if len(text.strip()) < 120:
                add(errors, f"{rel} is unexpectedly short")

def check_links(errors):
    link_re = re.compile(r"\[[^\]]+\]\(([^)]+)\)")
    for base in [ROOT, ROOT / "docs", ROOT / "templates"]:
        for path in base.rglob("*.md"):
            if ".git" in path.parts or ".work" in path.parts:
                continue
            text = path.read_text(encoding="utf-8", errors="ignore")
            for target in link_re.findall(text):
                if target.startswith(("http://","https://","mailto:","#")):
                    continue
                clean = target.split("#",1)[0]
                if not clean:
                    continue
                resolved = (path.parent / clean).resolve()
                try:
                    resolved.relative_to(ROOT)
                except ValueError:
                    add(errors, f"{path.relative_to(ROOT)} link escapes repo: {target}")
                    continue
                if not resolved.exists():
                    add(errors, f"{path.relative_to(ROOT)} broken link: {target}")

def check_inventory(errors):
    inv = ROOT / "archive-register/source-inventory.csv"
    if not inv.exists():
        add(errors, "source inventory missing")
        return
    rows = list(csv.DictReader(inv.open(encoding="utf-8")))
    if not rows:
        add(errors, "source inventory empty")
    ids = [r["inventory_id"] for r in rows]
    if len(ids) != len(set(ids)):
        add(errors, "duplicate inventory ids")
    for row in rows:
        if not re.fullmatch(r"[0-9a-f]{64}", row.get("sha256","")):
            add(errors, f"invalid checksum {row.get('inventory_id')}")

def check_spreadsheets(errors):
    paths = list((ROOT / "templates").rglob("*.xlsx"))
    try:
        from openpyxl import load_workbook
        for path in paths:
            wb = load_workbook(path, data_only=False)
            if not any(s in wb.sheetnames for s in ["Instructions", "1. Settings & Leaders", "1. Settings & Parameters", "1. Roster & Roles", "Roster & Roles", "Settings"]):
                add(errors, f"{path.relative_to(ROOT)} missing Instructions or Settings sheet")
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.startswith("=") and "#REF!" in cell.value:
                            add(errors, f"{path.relative_to(ROOT)} formula error {cell.coordinate}")
        return
    except Exception:
        pass
    import zipfile
    for path in paths:
        try:
            with zipfile.ZipFile(path) as z:
                names = z.namelist()
                if "xl/workbook.xml" not in names:
                    add(errors, f"{path.relative_to(ROOT)} missing workbook.xml")
                    continue
                workbook = z.read("xl/workbook.xml").decode("utf-8", "ignore")
                if not any(k in workbook for k in ['name="Instructions"', 'name="1. Settings &amp; Leaders"', 'name="1. Settings & Leaders"', 'name="1. Settings &amp; Parameters"', 'name="1. Roster &amp; Roles"', 'name="1. Roster & Roles"', 'name="Roster &amp; Roles"', 'name="Roster & Roles"', 'name="Settings"']):
                    add(errors, f"{path.relative_to(ROOT)} missing Instructions or Settings sheet")
                for name in names:
                    if name.startswith("xl/worksheets/") and name.endswith(".xml"):
                        xml = z.read(name).decode("utf-8", "ignore")
                        if "#REF!" in xml:
                            add(errors, f"{path.relative_to(ROOT)} formula reference error in {name}")
        except Exception as exc:
            add(errors, f"{path.relative_to(ROOT)} could not be inspected as xlsx: {exc}")

def check_safety_label(errors):
    required = "Draft requiring review by qualified safeguarding, legal, insurance and medical professionals before operational adoption."
    for path in list((ROOT / "docs/14-safety-and-safeguarding").glob("*.md")) + list((ROOT / "templates/safety").glob("*.md")):
        if required not in path.read_text(encoding="utf-8"):
            add(errors, f"{path.relative_to(ROOT)} missing safety draft label")

def check_site(errors):
    cfg = ROOT / "mkdocs.yml"
    if not cfg.exists():
        add(errors, "mkdocs.yml missing")
        return
    text = cfg.read_text(encoding="utf-8")
    try:
        import yaml
        data = yaml.safe_load(text) or {}
        nav_items = data.get("nav", [])
    except ImportError:
        # Fallback simple line-by-line regex parser for nav entries in mkdocs.yml
        nav_targets = re.findall(r":\s*([a-zA-Z0-9_\-/\.]+\.md)", text)
        for target in nav_targets:
            if not (ROOT / "docs" / target).exists():
                add(errors, f"mkdocs nav target missing: {target}")
        return

    for item in nav_items:
        if isinstance(item, dict):
            for _, target in item.items():
                if isinstance(target, str) and not (ROOT / "docs" / target).exists():
                    add(errors, f"mkdocs nav target missing: {target}")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--site-only", action="store_true")
    args = parser.parse_args()
    errors = []
    if args.site_only:
        check_site(errors)
    else:
        files = tracked_files()
        check_tracked(errors, files)
        check_privacy(errors, files)
        check_front_matter(errors)
        check_links(errors)
        check_inventory(errors)
        check_spreadsheets(errors)
        check_safety_label(errors)
        check_site(errors)
    if errors:
        print("\n".join(errors))
        raise SystemExit(1)
    print("validation passed")

if __name__ == "__main__":
    main()
