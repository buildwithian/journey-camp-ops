#!/usr/bin/env python3
from pathlib import Path
import sys
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
errors = []
for path in (ROOT / "templates").rglob("*.xlsx"):
    wb = load_workbook(path, data_only=False)
    if "Instructions" not in wb.sheetnames:
        errors.append(f"{path.relative_to(ROOT)} missing Instructions sheet")
    for ws in wb.worksheets:
        if ws.max_row > 5000:
            errors.append(f"{path.relative_to(ROOT)} has excessive rows in {ws.title}")
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and "#REF!" in cell.value:
                    errors.append(f"{path.relative_to(ROOT)} formula error in {cell.coordinate}")
if errors:
    print("\n".join(errors))
    raise SystemExit(1)
print("spreadsheet tests passed")
